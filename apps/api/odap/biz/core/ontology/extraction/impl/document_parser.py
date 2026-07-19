import logging
import os
from typing import Any, Dict, List, Optional

from odap.biz.core.ontology.extraction.interfaces.extraction_interfaces import DocumentParserInterface

logger = logging.getLogger(__name__)


class DocumentParser(DocumentParserInterface):
    def parse(self, file_path: str) -> str:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        ext = os.path.splitext(file_path)[1].lower()
        parsers = {
            ".pdf": self._parse_pdf,
            ".docx": self._parse_docx,
            ".doc": self._parse_docx,
            ".txt": self._parse_txt,
            ".md": self._parse_txt,
            ".csv": self._parse_csv,
            ".xlsx": self._parse_excel,
            ".xls": self._parse_excel,
            ".json": self._parse_json,
            ".xml": self._parse_xml,
            ".jpg": self._parse_image,
            ".jpeg": self._parse_image,
            ".png": self._parse_image,
            ".tiff": self._parse_image,
            ".tif": self._parse_image,
        }
        parser = parsers.get(ext)
        if not parser:
            raise ValueError(f"Unsupported file format: {ext}")
        return parser(file_path)

    def chunk_text(self, text: str, max_tokens: int = 4000) -> List[str]:
        chars_per_token = 2
        max_chars = max_tokens * chars_per_token
        if len(text) <= max_chars:
            return [text]
        chunks = []
        start = 0
        while start < len(text):
            end = start + max_chars
            if end < len(text):
                paragraph_break = text.rfind("\n\n", start, end)
                if paragraph_break > start:
                    end = paragraph_break + 2
            chunks.append(text[start:end])
            start = end
        return chunks

    def _parse_pdf(self, file_path: str) -> str:
        try:
            from odap.biz.core.ontology.construction.ingestion.services import PDFProcessor
            processor = PDFProcessor()
            result = processor.extract_text(file_path)
            return result if isinstance(result, str) else str(result)
        except Exception as e:
            logger.warning(f"PDF parsing failed, trying PyPDF2: {e}")
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(file_path)
                return "\n".join(page.extract_text() or "" for page in reader.pages)
            except Exception as e2:
                raise ValueError(f"PDF parsing failed: {e2}")

    def _parse_docx(self, file_path: str) -> str:
        try:
            from odap.biz.core.ontology.construction.ingestion.services import WordProcessor
            processor = WordProcessor()
            result = processor.extract_text(file_path)
            return result if isinstance(result, str) else str(result)
        except Exception as e:
            logger.warning(f"Word parsing failed, trying python-docx: {e}")
            try:
                from docx import Document
                doc = Document(file_path)
                return "\n".join(p.text for p in doc.paragraphs)
            except Exception as e2:
                raise ValueError(f"Word parsing failed: {e2}")

    def _parse_txt(self, file_path: str) -> str:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()

    def _parse_csv(self, file_path: str) -> str:
        import csv
        rows = []
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(", ".join(row))
        return "\n".join(rows)

    def _parse_excel(self, file_path: str) -> str:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            parts = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    parts.append(", ".join(str(c) for c in row if c is not None))
            wb.close()
            return "\n".join(parts)
        except Exception as e:
            raise ValueError(f"Excel parsing failed: {e}")

    def _parse_json(self, file_path: str) -> str:
        import json
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return json.dumps(data, ensure_ascii=False, indent=2)

    def _parse_xml(self, file_path: str) -> str:
        import xml.etree.ElementTree as ET
        tree = ET.parse(file_path)
        root = tree.getroot()
        parts = []
        for elem in root.iter():
            if elem.text and elem.text.strip():
                parts.append(f"{elem.tag}: {elem.text.strip()}")
        return "\n".join(parts)

    def _parse_image(self, file_path: str) -> str:
        try:
            from odap.biz.core.ontology.construction.ingestion.services import OCRProcessor
            processor = OCRProcessor()
            result = processor.extract_text(file_path)
            return result if isinstance(result, str) else str(result)
        except Exception as e:
            logger.warning(f"OCR parsing failed, trying pytesseract: {e}")
            try:
                import pytesseract
                from PIL import Image
                img = Image.open(file_path)
                return pytesseract.image_to_string(img, lang="chi_sim+eng")
            except Exception as e2:
                raise ValueError(f"Image OCR failed: {e2}")
